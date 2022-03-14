import datetime
from typing import List

from openpyxl import Workbook
from pydantic import BaseModel, validator
from rdflib import Graph, URIRef, Literal
from rdflib.namespace import DCAT, DCTERMS, OWL, SKOS, RDF, RDFS, XSD

try:
    from utils import all_strings_in_list_are_iris, string_is_http_iri
except:
    import sys

    sys.path.append("..")
    from vocexcel.utils import all_strings_in_list_are_iris, string_is_http_iri

ORGANISATIONS = {
    "CGI": URIRef("https://linked.data.gov.au/org/cgi"),
    "GA": URIRef("https://linked.data.gov.au/org/ga"),
    "GGIC": URIRef("https://linked.data.gov.au/org/ggic"),
    "GSQ": URIRef("https://linked.data.gov.au/org/gsq"),
    "ICSM": URIRef("https://linked.data.gov.au/org/icsm"),
    "DES": URIRef("https://linked.data.gov.au/org/des"),
    "BITRE": URIRef("https://linked.data.gov.au/org/bitre"),
    "CASA": URIRef("https://linked.data.gov.au/org/casa"),
    "ATSB": URIRef("https://linked.data.gov.au/org/atsb"),
}

ORGANISATIONS_INVERSE = {
    URIRef("https://linked.data.gov.au/org/cgi"): "CGI",
    URIRef("https://linked.data.gov.au/org/ga"): "GA",
    URIRef("https://linked.data.gov.au/org/ggic"): "GGIC",
    URIRef("https://linked.data.gov.au/org/gsq"): "GSQ",
    URIRef("https://linked.data.gov.au/org/icsm"): "ICSM",
    URIRef("https://linked.data.gov.au/org/des"): "DES",
    URIRef("https://linked.data.gov.au/org/bitre"): "BITRE",
    URIRef("https://linked.data.gov.au/org/casa"): "CASA",
    URIRef("https://linked.data.gov.au/org/atsb"): "ATSB",
}


class ConceptScheme(BaseModel):
    uri: str
    title: str
    description: str
    created: datetime.date
    modified: datetime.date = None
    creator: str
    publisher: str
    provenance: str
    version: str = None
    custodian: str = None
    pid: str = None

    @validator("creator")
    def creator_must_be_from_list(cls, v):
        if v not in ORGANISATIONS.keys():
            raise ValueError(
                f"Organisations must selected from the Organisations list: {', '.join(ORGANISATIONS)}"
            )
        return v

    @validator("publisher")
    def publisher_must_be_from_list(cls, v):
        if v not in ORGANISATIONS.keys():
            raise ValueError(
                f"Organisations must selected from the Organisations list: {', '.join(ORGANISATIONS)}"
            )
        return v

    def to_graph(self):
        g = Graph()
        v = URIRef(self.uri)
        g.add((v, RDF.type, SKOS.ConceptScheme))
        g.add((v, SKOS.prefLabel, Literal(self.title, lang="en")))
        g.add((v, SKOS.definition, Literal(self.description, lang="en")))
        g.add((v, DCTERMS.created, Literal(self.created, datatype=XSD.date)))
        if self.modified is not None:
            g.add((v, DCTERMS.modified, Literal(self.created, datatype=XSD.date)))
        else:
            g.add(
                (
                    v,
                    DCTERMS.modified,
                    Literal(
                        datetime.datetime.now().strftime("%Y-%m-%d"), datatype=XSD.date
                    ),
                )
            )
        g.add((v, DCTERMS.creator, ORGANISATIONS[self.creator]))
        g.add((v, DCTERMS.publisher, ORGANISATIONS[self.publisher]))
        if self.version is not None:
            g.add((v, OWL.versionInfo, Literal(self.version)))
        g.add((v, DCTERMS.provenance, Literal(self.provenance, lang="en")))
        if self.custodian is not None:
            g.add((v, DCAT.contactPoint, Literal(self.custodian)))
        if self.pid is not None:
            # adding to the graph depending on if the pid is a URI or a literal
            if string_is_http_iri(self.pid)[0]:
                g.add((v, RDFS.seeAlso, URIRef(self.pid)))
            else:
                g.add((v, RDFS.seeAlso, Literal(self.pid)))

        # bind non-core prefixes
        g.bind("cs", v)
        g.bind(
            "",
            str(v).split("#")[0] if "#" in str(v) else "/".join(str(v).split("/")[:-1]),
        )
        g.bind("dcat", DCAT)
        g.bind("dcterms", DCTERMS)
        g.bind("skos", SKOS)
        g.bind("owl", OWL)

        return g

    def to_excel(self, wb: Workbook):
        ws = wb.active
        ws["B1"] = self.uri
        ws["B2"] = self.title
        ws["B3"] = self.description
        ws["B4"] = self.created.isoformat()
        ws["B5"] = self.modified.isoformat()
        ws["B6"] = self.creator
        ws["B7"] = self.publisher
        ws["B8"] = self.version
        ws["B9"] = self.provenance
        # ws["B10"] = ""
        # ws["B11"] = ""


class Concept(BaseModel):
    uri: str
    pref_label: str
    alt_labels: List[str] = []
    pl_language_code: List[str] = []
    definition: str
    def_language_code: List[str] = []
    children: List[str] = []
    other_ids: List[str] = []
    home_vocab_uri: str = None
    provenance: str = None
    related_match: List[str] = []
    close_match: List[str] = []
    exact_match: List[str] = []
    narrow_match: List[str] = []
    broad_match: List[str] = []

    @validator("children")
    def each_child_must_be_an_iri(cls, elem):
        r = all_strings_in_list_are_iris(elem)
        assert r[0], r[1]
        return elem

    @validator("related_match")
    def each_rm_must_be_an_iri(cls, elem):
        r = all_strings_in_list_are_iris(elem)
        assert r[0], r[1]
        return elem

    @validator("close_match")
    def each_cm_must_be_an_iri(cls, elem):
        r = all_strings_in_list_are_iris(elem)
        assert r[0], r[1]
        return elem

    @validator("exact_match")
    def each_em_must_be_an_iri(cls, elem):
        r = all_strings_in_list_are_iris(elem)
        assert r[0], r[1]
        return elem

    @validator("narrow_match")
    def each_nm_must_be_an_iri(cls, elem):
        r = all_strings_in_list_are_iris(elem)
        assert r[0], r[1]
        return elem

    @validator("broad_match")
    def each_bm_must_be_an_iri(cls, elem):
        r = all_strings_in_list_are_iris(elem)
        assert r[0], r[1]
        return elem

    def to_graph(self):
        g = Graph()
        c = URIRef(self.uri)
        g.add((c, RDF.type, SKOS.Concept))
        if not self.pl_language_code:
            g.add((c, SKOS.prefLabel, Literal(self.definition, lang="en")))
        else:
            for lang_code in self.pl_language_code:
                g.add((c, SKOS.prefLabel, Literal(self.definition, lang=lang_code)))
        if self.alt_labels is not None:
            for alt_label in self.alt_labels:
                g.add((c, SKOS.altLabel, Literal(alt_label, lang="en")))
        if not self.def_language_code:
            g.add((c, SKOS.definition, Literal(self.definition, lang="en")))
        else:
            for lang_code in self.def_language_code:
                g.add((c, SKOS.definition, Literal(self.definition, lang=lang_code)))
        for child in self.children:
            g.add((c, SKOS.narrower, URIRef(child)))
            g.add((URIRef(child), SKOS.broader, c))
        if self.other_ids is not None:
            for other_id in self.other_ids:
                g.add((c, SKOS.notation, Literal(other_id)))
        if self.home_vocab_uri is not None:
            g.add((c, RDFS.isDefinedBy, URIRef(self.home_vocab_uri)))
        # if self.provenance is not None:
        #     g.add((c, DCTERMS.provenance, Literal(self.provenance, lang="en")))
        if self.related_match is not None:
            for related_match in self.related_match:
                g.add((c, SKOS.relatedMatch, URIRef(related_match)))
        if self.close_match:
            for close_match in self.close_match:
                g.add((c, SKOS.closeMatch, URIRef(close_match)))
        if self.exact_match is not None:
            for exact_match in self.exact_match:
                g.add((c, SKOS.exactMatch, URIRef(exact_match)))
        if self.narrow_match is not None:
            for narrow_match in self.narrow_match:
                g.add((c, SKOS.narrowMatch, URIRef(narrow_match)))
        if self.broad_match is not None:
            for broad_match in self.broad_match:
                g.add((c, SKOS.broadMatch, URIRef(broad_match)))

        return g

    def to_excel(self, wb: Workbook, row_no: int):
        ws = wb.active
        ws[f"A{row_no}"] = self.uri
        ws[f"B{row_no}"] = self.pref_label
        ws[f"C{row_no}"] = self.alt_labels
        ws[f"D{row_no}"] = self.definition
        ws[f"E{row_no}"] = ",\n".join(self.children)
        ws[f"F{row_no}"] = ",\n".join(self.other_ids)
        ws[f"G{row_no}"] = self.home_vocab_uri
        ws[f"H{row_no}"] = self.provenance


class Collection(BaseModel):
    uri: str
    pref_label: str
    definition: str
    members: List[str]
    provenance: str = None

    @validator("members")
    def members_must_by_iris(cls, v):
        if not v[0].startswith("http"):
            raise ValueError("The members of a Collection must be a list of IRIs")
        return v

    def to_graph(self):
        g = Graph()
        c = URIRef(self.uri)
        g.add((c, RDF.type, SKOS.Collection))
        g.add((c, SKOS.prefLabel, Literal(self.pref_label, lang="en")))
        g.add((c, SKOS.definition, Literal(self.definition, lang="en")))
        for member in self.members:
            g.add((c, SKOS.member, URIRef(member)))
        if self.provenance is not None:
            g.add((c, DCTERMS.provenance, Literal(self.provenance, lang="en")))

        return g

    def to_excel(self, wb: Workbook, row_no: int):
        ws = wb.active
        ws[f"A{row_no}"] = self.uri
        ws[f"B{row_no}"] = self.pref_label
        ws[f"C{row_no}"] = self.definition
        ws[f"D{row_no}"] = ",\n".join(self.members)
        ws[f"E{row_no}"] = self.provenance


class Vocabulary(BaseModel):
    concept_scheme: ConceptScheme
    concepts: List[Concept]
    collections: List[Collection]

    def to_graph(self):
        g = self.concept_scheme.to_graph()
        cs = URIRef(self.concept_scheme.uri)
        for concept in self.concepts:
            g += concept.to_graph()
            g.add((URIRef(concept.uri), SKOS.inScheme, cs))
        for collection in self.collections:
            g += collection.to_graph()
            g.add((URIRef(collection.uri), DCTERMS.isPartOf, cs))
            g.add((cs, DCTERMS.hasPart, URIRef(collection.uri)))

        # create as Top Concepts those Concepts that have no skos:narrower properties with them as objects
        for s in g.subjects(SKOS.inScheme, cs):
            is_tc = True
            for _ in g.objects(s, SKOS.broader):
                is_tc = False
            if is_tc:
                g.add((cs, SKOS.hasTopConcept, s))
                g.add((s, SKOS.topConceptOf, cs))

        return g
