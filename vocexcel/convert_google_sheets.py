import gspread
from oauth2client.service_account import ServiceAccountCredentials

from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

try:
    import models
    from utils import split_and_tidy, ConversionError, load_workbook
except:
    import sys

    sys.path.append("..")
    from vocexcel import models
    from vocexcel.utils import split_and_tidy, ConversionError, load_workbook

scope = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/drive",
]
creds = ServiceAccountCredentials.from_json_keyfile_name("creds.json", scope)
client = gspread.authorize(creds)

concept_sheet = client.open("Google Sheets test VocExcel-template_043").worksheet(
    "Concepts*"
)
concept_scheme_sheet = client.open(
    "Google Sheets test VocExcel-template_043"
).worksheet("Concept Scheme*")
additional_concept_sheet = client.open(
    "Google Sheets test VocExcel-template_043"
).worksheet("Additional Concept " "Features")
collections_sheet = client.open("Google Sheets test VocExcel-template_043").worksheet(
    "Collections"
)
prefix_sheet = client.open("Google Sheets test VocExcel-template_043").worksheet(
    "Prefix Sheet"
)


def split_and_tidy_gsheets(cell_value: str):
    # with the new logic -> I need the extra line of logic regarding cell_value
    if cell_value == "" or cell_value is None:
        return []
    return (
        [x.strip() for x in cell_value.strip().split(",")]
        if cell_value is not None
        else []
    )


def create_prefix_dict_gsheets(s: gspread.Worksheet):
    prefix_dict = {}
    prefix_sheet_values = s.get_values()
    row = 2
    for values in prefix_sheet_values:
        if values[0] == "Prefix":
            pass
        else:
            try:
                prefix_dict[values[0]] = values[1]
                row += 1
            except Exception as e:
                raise ConversionError(
                    f"Prefix Processing error, sheet {s}, row {row}, error {e}"
                )
    return prefix_dict


def only_one_colon_in_str(string):
    counter = 0
    for character in string:
        if character == ":":
            counter += 1
    if counter != 1:
        return False
    return True


def using_prefix_non_list_output(
    cell_value, prefix: dict, s: gspread.Worksheet, row: int
):
    c = cell_value
    if c is not None:
        if c.startswith("http"):
            pass
        elif only_one_colon_in_str(c):
            split_c_prefix = c.split(":")[0]
            split_c_added_component = c.split(":")[1]

            if split_c_prefix in prefix:
                c = prefix[split_c_prefix] + split_c_added_component
            else:
                print(
                    f"The prefix used: '{split_c_prefix}' in sheet: {s} and row {row} "
                    f" isn't included in the prefix sheet"
                )
                raise Exception
        else:
            print(
                f"Something doesn't look right on sheet: {s} and row {row}."
                f"Likely this isn't in http form or more colons are used than normal"
            )
            raise Exception
    return c


def using_prefix_list_output(cell_value, prefix: dict, s: gspread.Worksheet, row: int):
    variables = []
    for i in cell_value:
        if cell_value is not None:
            try:
                c = i
                if c.startswith("http"):
                    pass
                elif only_one_colon_in_str(c):
                    split_c_prefix = c.split(":")[0]
                    split_c_added_component = c.split(":")[1]

                    if split_c_prefix in prefix:
                        c = prefix[split_c_prefix] + split_c_added_component
                    else:
                        print(
                            f"The prefix used: {split_c_prefix} in sheet {s} and row {row} "
                            f"isn't included in the prefix sheet"
                        )
                        raise Exception
                else:
                    print(
                        f"Something doesn't look right on row {row} and sheet {s}. "
                        f"Likely this isn't in http form or more colons are used than normal"
                    )
                    raise Exception
                variables.append(c)
            except Exception as e:
                print(e)
    return variables


def extract_concept_scheme(sheet: Worksheet, prefix: dict):
    cs = models.ConceptScheme(
        uri=using_prefix_non_list_output(sheet.cell(2, 2).value, prefix, sheet, 2),
        title=sheet.cell(3, 2).value,
        description=sheet.cell(4, 2).value,
        created=sheet.cell(5, 2).value,
        modified=sheet.cell(6, 2).value,
        creator=sheet.cell(7, 2).value,
        publisher=sheet.cell(8, 2).value,
        version=sheet.cell(9, 2).value,
        provenance=sheet.cell(10, 2).value,
        custodian=sheet.cell(11, 2).value,
        pid=sheet.cell(12, 2).value,
    )
    return cs


def extract_concept(
    csheet: gspread.Worksheet, acsheet: gspread.Worksheet, prefix: dict
):
    all_concept_sheet_values = csheet.get_values()
    ac_val = acsheet.get_values()
    concepts = []
    row = 3
    for rows in all_concept_sheet_values:
        if rows[0] == "Concepts" or rows[0] == "Concept IRI*":
            pass
        else:
            try:
                c = models.Concept(
                    uri=using_prefix_non_list_output(rows[0], prefix, csheet, row),
                    pref_label=rows[1],
                    pl_language_code=split_and_tidy_gsheets(rows[2]),
                    definition=rows[3],
                    def_language_code=split_and_tidy_gsheets(rows[4]),
                    alt_labels=split_and_tidy_gsheets(rows[5]),
                    children=split_and_tidy_gsheets(rows[6]),
                    provenance=rows[7],
                    home_vocab_uri=using_prefix_non_list_output(
                        rows[8], prefix, csheet, row
                    ),
                    # additional concepts sheet page
                    related_match=using_prefix_list_output(
                        split_and_tidy_gsheets(ac_val[row - 1][1]), prefix, acsheet, row
                    ),
                    close_match=using_prefix_list_output(
                        split_and_tidy_gsheets(ac_val[row - 1][2]), prefix, acsheet, row
                    ),
                    exact_match=using_prefix_list_output(
                        split_and_tidy_gsheets(ac_val[row - 1][3]), prefix, acsheet, row
                    ),
                    narrow_match=using_prefix_list_output(
                        split_and_tidy_gsheets(ac_val[row - 1][4]), prefix, acsheet, row
                    ),
                    broad_match=using_prefix_list_output(
                        split_and_tidy_gsheets(ac_val[row - 1][5]), prefix, acsheet, row
                    ),
                )
                row += 1
                concepts.append(c)
            except ValidationError as e:
                raise ConversionError(
                    f" Concept Processing error likely at sheet {csheet} or {acsheet}, row {row}, and has error {e}"
                )
    return concepts


def extract_collection(col_sheet: gspread.Worksheet, prefix: dict):
    collection_values = col_sheet.get_values()
    collections = []
    row = 3
    for rows in collection_values:
        if rows[0] == "Collections" or rows[0] == "Collection URI":
            pass
        else:
            try:
                c = models.Collection(
                    uri=using_prefix_non_list_output(rows[0], prefix, col_sheet, row),
                    pref_label=rows[1],
                    definition=rows[2],
                    members=using_prefix_list_output(
                        split_and_tidy_gsheets(rows[3]), prefix, col_sheet, row
                    ),
                    provenance=rows[4],
                )
                collections.append(c)
            except ValidationError as e:
                raise ConversionError(
                    f" Collection Processing Error likely at {col_sheet}, row {row}, and has error {e}"
                )
    return collections


if __name__ == "__main__":
    prefix = create_prefix_dict_gsheets(prefix_sheet)
    print(extract_collection(collections_sheet, prefix))
    print(extract_concept(concept_sheet, additional_concept_sheet, prefix))
    print(extract_concept_scheme(concept_scheme_sheet, prefix))
