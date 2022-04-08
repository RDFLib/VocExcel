from typing import Tuple, List

from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

try:
    import models
    from utils import split_and_tidy, ConversionError, load_workbook
except:
    import sys

    sys.path.append("..")
    from vocexcel import models
    from vocexcel.utils import split_and_tidy, ConversionError, load_workbook


# Checking how many colon's in string
def only_one_colon_in_str(string):
    counter = 0
    for character in string:
        if character == ':':
            counter += 1
    if counter != 1:
        return False
    return True


# function in creating a dictionary / recalling from a dictionary.
def create_prefix_dict(s: Worksheet):
    # create an empty dict
    prefix_dict = {}

    # add prefix values according to the prefix sheet
    for col in s.iter_cols(max_col=1):
        for cell in col:
            row = cell.row
            if (
                    cell.value is None
                    or cell.value == "Prefix"
                    or cell.value == "Prefix Sheet"
            ):
                pass
            else:
                # dynamically allocate the prefix sheet
                try:
                    prefix_dict[s[f"A{row}"].value] = s[f"B{row}"].value
                except Exception as e:
                    raise ConversionError(
                        f"Prefix processing error, sheet {s}, row {row}, error: {e}"
                    )
    return prefix_dict


# prefix and namespace use without list output
def using_prefix_and_namespace_non_list_output(cell_value, prefix, s: Worksheet, row):
    c = cell_value
    if c is not None:
        if c.startswith('http://') or c.startswith('https://'):
            pass
        elif only_one_colon_in_str(c):
            split_c_prefix = c.split(":")[0]
            split_c_added_component = c.split(":")[1]

            if split_c_prefix in prefix:
                c = prefix[split_c_prefix] + split_c_added_component
            else:
                print(
                    f"the prefix used: '{split_c_prefix}' in sheet {s} and row {row} "
                    f"isn't included in the prefix sheet")
                raise Exception
        else:
            print(
                f"Something doesn't look right on row {row} and sheet {s}. "
                f"Likely this isn't in http form or more colons are used than normal")
            raise Exception
    return c


# specifically just for concept scheme
def using_prefix_and_namespace_cs(cell_value, prefix):
    c = cell_value
    if c is not None:
        if c.startswith('http://') or c.startswith('https://'):
            pass
        elif only_one_colon_in_str(c):
            split_c_prefix = c.split(":")[0]
            split_c_added_component = c.split(":")[1]

            if split_c_prefix in prefix:
                c = prefix[split_c_prefix] + split_c_added_component
            else:
                print(
                    f"the prefix used: '{split_c_prefix}' in the concept scheme page"
                    f"isn't included in the prefix sheet")
                raise Exception
        else:
            print(
                f"Something doesn't look right in the concept scheme page. "
                f"Likely the cells using prefixes aren't in http form or more colons are used than normal")
            raise Exception
    return c


# prefix and namespace list with list output
def using_prefix_and_namespace(cell_value, prefix, s: Worksheet, row):
    variables = []
    for i in cell_value:
        if cell_value is not None:
            try:
                c = i
                if c.startswith('http://') or c.startswith('https://'):
                    pass
                elif only_one_colon_in_str(c):
                    split_c_prefix = c.split(":")[0]
                    split_c_added_component = c.split(":")[1]

                    if split_c_prefix in prefix:
                        c = prefix[split_c_prefix] + split_c_added_component
                    else:
                        print(
                            f"the prefix used: '{split_c_prefix}' in sheet {s} and row {row} "
                            f"isn't included in the prefix sheet")
                        raise Exception
                else:
                    print(
                        f"Something doesn't look right on row {row} and sheet {s}. "
                        f"Likely this isn't in http form or more colons are used than normal")
                    raise Exception
                variables.append(c)
            except Exception as e:
                print(e)
    return variables


# template version 0.4.3 using prefixes
def extract_concepts_and_collections(
        q: Worksheet, r: Worksheet, s: Worksheet, prefix
) -> Tuple[List[models.Concept], List[models.Collection]]:
    concepts = []
    collections = []
    # Iterating over the concept page and the additional concept page
    for col in q.iter_cols(max_col=1):
        for cell in col:
            row = cell.row
            if (
                    cell.value is None
                    or cell.value == "Concepts"
                    or cell.value == "Concept IRI*"
            ):
                pass
            else:
                try:
                    c = models.Concept(
                        uri=using_prefix_and_namespace_non_list_output(q[f"A{row}"].value, prefix, q, row),
                        pref_label=q[f"B{row}"].value,
                        pl_language_code=split_and_tidy(q[f"C{row}"].value),
                        definition=q[f"D{row}"].value,
                        def_language_code=split_and_tidy(q[f"E{row}"].value),
                        alt_labels=split_and_tidy(q[f"F{row}"].value),
                        children=using_prefix_and_namespace(split_and_tidy(q[f"G{row}"].value), prefix, q, row),
                        provenance=q[f"H{row}"].value,
                        # Note in the new template, home_vocab_uri is synonymous with source vocab uri
                        home_vocab_uri=using_prefix_and_namespace_non_list_output(q[f"I{row}"].value, prefix, q, row),
                        # additional concept features sheets
                        related_match=using_prefix_and_namespace(split_and_tidy(r[f"B{row}"].value), prefix, r, row),
                        close_match=using_prefix_and_namespace(split_and_tidy(r[f"C{row}"].value), prefix, r, row),
                        exact_match=using_prefix_and_namespace(split_and_tidy(r[f"D{row}"].value), prefix, r, row),
                        narrow_match=using_prefix_and_namespace(split_and_tidy(r[f"E{row}"].value), prefix, r, row),
                        broad_match=using_prefix_and_namespace(split_and_tidy(r[f"F{row}"].value), prefix, r, row),
                    )
                    concepts.append(c)
                except ValidationError as e:
                    raise ConversionError(
                        f"Concept processing error likely at sheet {q}, column {col}, row {row}, and has error: {e}"
                    )

    # iterating over the collections page
    for col in s.iter_cols(max_col=1):
        for cell in col:
            row = cell.row
            if (
                    cell.value is None
                    or cell.value == "Collections"
                    or cell.value == "Collection URI"
            ):
                pass
            else:
                try:
                    c = models.Collection(
                        uri=using_prefix_and_namespace_cs(s[f"A{row}"].value, prefix),
                        pref_label=s[f"B{row}"].value,
                        definition=s[f"C{row}"].value,
                        members=using_prefix_and_namespace(split_and_tidy(s[f"D{row}"].value), prefix, s, row),
                        provenance=s[f"E{row}"].value,
                    )
                    collections.append(c)
                except ValidationError as e:
                    raise ConversionError(
                        f"Collection processing error, likely at sheet {s}, column {col}, row {row}, and has error: {e}"
                    )
    return concepts, collections


def extract_concept_scheme(sheet: Worksheet, prefix):
    cs = models.ConceptScheme(
        uri=using_prefix_and_namespace_cs(sheet["B2"].value, prefix),
        title=sheet["B3"].value,
        description=sheet["B4"].value,
        created=sheet["B5"].value,
        modified=sheet["B6"].value,
        creator=sheet["B7"].value,
        publisher=sheet["B8"].value,
        version=sheet["B9"].value,
        provenance=sheet["B10"].value,
        custodian=sheet["B11"].value,
        pid=sheet["B12"].value,
    )
    return cs
